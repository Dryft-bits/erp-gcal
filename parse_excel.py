from pathlib import Path

from openpyxl import load_workbook
from utils import config, read_json, to_title, write_json


def iter_rows(wb, column_map):
    """Generator for all the rows in the workbook"""
    for sheet in wb:
        rows = sheet.rows
        next(rows)
        next(rows)
        for row in rows:
            if not row:
                continue
            yield {name: row[col].value for name, col in column_map.items()}


def parse_main_tt(file_path: Path):
    COLUMNS = {  # 0 indexed column indices, The indices need to be changed every according to the excel
        "c_num": 1,
        "c_title": 2,
        "sec_num": 6,
        "instr_name": 7,
        "room": 8,
        "days": 9,
        "hours": 11,
        "midsem": 12,
        "compre": 13,
    }
    workbook = load_workbook(file_path, read_only=True)
    course_db = {}
    for data in iter_rows(workbook, COLUMNS):
        if not data["instr_name"]:
            continue  # blank row
        if data["c_num"] == "COURSE NO":
            continue
        if data["sec_num"] == "S E C":
            continue

        # new Course
        if data["c_num"]:
            course = {
                "name": to_title(data["c_title"]),
                "sections": {},
            }
            if data["compre"]:
                date, sess = str(data["compre"]).split()
                course["compre"] = {"date": date, "session": sess}

            if data["midsem"]:
                if data["midsem"] == "TBA":
                    course["midsem"] = {"date": "TBA", "time": "TBA"}
                else:
                    try:
                        date, start_time, end_time = data["midsem"].split()
                        # print("correct:", data["midsem"].split())
                        course["midsem"] = {
                            "date": str(date.strip()),
                            "time": start_time + "-" + end_time,
                        }
                    except:
                        if "-" in data["midsem"]:
                            data["midsem"] = data["midsem"].replace("-", "")
                            # print("edited:", data["midsem"].split())
                            date, start_time, end_time = data["midsem"].split()
                            course["midsem"] = {
                                "date": str(date.strip()),
                                "time": start_time + "-" + end_time,
                            }
                        else:
                            print("Error still not fixed...")

            course_db[data["c_num"]] = course  # add to course
            sec_type = "L"
            sec_num_counter = 1

        # new Tutorial or Practical section
        if not data["c_num"] and data["c_title"]:
            sec_type = data["c_title"][0]
            sec_num_counter = 1

        # new Section
        if (
            data["instr_name"]
            and data.get("room", True)
            and not sec_type == "L"
            or data["sec_num"]
        ) or data["c_title"]:
            sec_num = int(data["sec_num"] or sec_num_counter)
            section = {"instructors": [], "sched": []}
            course["sections"][sec_type + str(sec_num)] = section
            sec_num_counter += 1
            instructors = set()  # keep track of unique instructors

        if isinstance(data.get("hours"), (float, int)):
            data["hours"] = str(int(data["hours"]))
        if data.get("days"):
            hours = []
            for hour in map(int, data["hours"].split()):
                if hour > 15:
                    hours.extend(map(int, str(hour)))
                else:
                    hours.append(hour)
            days = data["days"].split()
            sched = {"room": data.get("room", "NA"), "days": days}
            if len(hours) == hours[-1] - hours[0] + 1:  # continuous hours
                section["sched"].append(dict(**sched, hours=hours))
            else:
                for hour in hours:  # separate sched for each hour
                    section["sched"].append(dict(**sched, hours=(hour,)))
        if data["instr_name"].lower() not in instructors:
            section["instructors"].append(data["instr_name"])
            instructors.add(data["instr_name"].lower())
    return course_db


def parse_files(tt_file: Path, midsem_file: Path):
    timetable = parse_main_tt(tt_file)
    return timetable


class CourseDB:
    tt_file = Path(config["COURSES"]["tt_file"])
    midsem_file = Path(config["COURSES"].get("midsem_file", ""))
    _course_db = None
    _timetable = None

    def __new__(cls, *args, **kwargs):
        if not cls._course_db:
            cls._course_db = super().__new__(cls, *args, **kwargs)
        return cls._course_db

    def get_timetable(self, force_parse=False):
        if not self.tt_file.exists():
            raise FileNotFoundError(self.tt_file)
        json_file = self.tt_file.with_suffix(".json")
        if not json_file.exists() or (force_parse and self.tt_file.suffix != ".json"):
            self._timetable = parse_files(self.tt_file, self.midsem_file)
            write_json(json_file, self._timetable)
        else:
            self._timetable = read_json(json_file)
        return self._timetable

    def __getitem__(self, course_code):
        if not self._timetable:
            self._timetable = self.get_timetable()
        return self._timetable.get(course_code)

    @property
    def timetable(self):
        if not self._timetable:
            self._timetable = self.get_timetable()
        return self._timetable


course_db = CourseDB()


def main():
    CourseDB().get_timetable(True)


if __name__ == "__main__":
    main()
